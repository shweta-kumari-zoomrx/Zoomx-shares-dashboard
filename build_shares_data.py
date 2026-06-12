"""
build_shares_data.py  --  Shares Dashboard data builder
========================================================
SELF-CONTAINED: reads ONLY the RD Excel for data + the local shares_maps.py for
survey-definition metadata (drug<->A-code maps, series, region map). No dependency
on the deck (generate_slides.py) or the PPTX. Same methodology as the deck:

  - GLOBAL_OL respondents dropped (61 remain)
  - Outlier union (ol_ids) computed from S0_80Z_A1/A2/A3 -> flagged, NOT used in shares
  - Current shares  = Q3_40Z_{E/O/C}_{IHC3/IHC2/IHCU}_{1L/2L/3L}_{Acode}
  - Future  shares  = Q6_10Z_{EC/OC/CC}_...        (the deck's slides 11-18 set)
  - Weights         = Q3_25Z_{E/O/C}_{1L/2L/3L}_{IHC3/IHC2/IHCU}

Outputs:
  public/Shares_RD.xlsx   one row per respondent (PII stripped): RespID, OL flag,
                          Practice Setting, Specialty, Region, + every numeric
                          Q3_40 / Q6_10 / Q3_25 column referenced by the maps.
  src/sharesConfig.json   series lists, knownMissing, and current/future column
                          maps so the browser can recompute shares live per filter.

Run from this folder:  python build_shares_data.py
"""
import os, re, json, sys
import openpyxl
import shares_maps as gen   # SELF-CONTAINED local metadata (no deck dependency)

HERE = os.path.dirname(os.path.abspath(__file__))
PARENT = os.path.dirname(HERE)
# Single data source: the RD Excel. Override with env var if it ever moves.
EXCEL_FILE = os.environ.get("SHARES_RD", os.path.join(PARENT, "ENH Gyn RD 2026.xlsx"))

IND_LETTER = {"EC": "E", "OC": "O", "CC": "C"}
STRATA = ["IHC3", "IHC2", "IHCU"]
LOTS = ["1L", "2L", "3L"]
INDS = ["EC", "OC", "CC"]


def future_to_current(col):
    """Q6_10Z_EC_IHC3_1L_A1 -> Q3_40Z_E_IHC3_1L_A1 ; ..._OTH3_1 -> ..._other"""
    if col is None:
        return None
    c = col.replace("Q6_10Z_", "Q3_40Z_")
    c = re.sub(r"^(Q3_40Z_)(EC|OC|CC)_",
               lambda m: m.group(1) + IND_LETTER[m.group(2)] + "_", c)
    c = re.sub(r"_OTH[32U]_\d+$", "_other", c)
    return c


def build_maps(future_map_dict):
    """Return {ind:{lot:{series:{'future':col,'current':col}}}} for one stratum."""
    out = {}
    for ind in INDS:
        out[ind] = {}
        for lot in LOTS:
            out[ind][lot] = {}
            for series, fcol in future_map_dict[ind].get(lot, {}).items():
                out[ind][lot][series] = {"future": fcol, "current": future_to_current(fcol)}
    return out


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[9])
    data = rows[10:]
    hidx = {name: i for i, name in enumerate(headers) if name}

    # drop GLOBAL_OL
    ot = hidx.get("Outlier Type")
    if ot is not None:
        data = [r for r in data if r[ot] != "GLOBAL_OL"]

    # ── outlier union, replicating generate_slides.ol_mean ────────────────────
    import statistics

    def outlier_fence(nums):
        q1 = statistics.quantiles(nums, n=4)[0]
        q3 = statistics.quantiles(nums, n=4)[2]
        iqr_fence = q3 + 1.5 * (q3 - q1)
        sd2_fence = statistics.mean(nums) + 2 * statistics.stdev(nums)
        return (iqr_fence + sd2_fence) / 2

    def ol_set(col_name):
        idx = hidx.get(col_name)
        if idx is None:
            return set()
        rv = [(i, r[idx]) for i, r in enumerate(data) if isinstance(r[idx], (int, float))]
        if len(rv) < 2:
            return set()
        nums = [v for _, v in rv]
        fence = outlier_fence(nums)
        return {i for i, v in rv if v > fence}

    ol_ids = ol_set("S0_80Z_A1") | ol_set("S0_80Z_A2") | ol_set("S0_80Z_A3")

    # ── segment helpers (mirror generate_slides) ──────────────────────────────
    def practice(v):
        s = str(v)
        if "Academic" in s or "University" in s: return "University / Academic"
        if "Cancer" in s or "Spec" in s:         return "Cancer / Specialized"
        if "Office" in s:                         return "Office-based"
        if "Community" in s or "General" in s:    return "Community / General"
        if "Private" in s:                        return "Private clinic"
        return "Other"

    def specialty(v):
        s = str(v)
        if "Medical" in s: return "Medical Oncologist"
        if "Gynecol" in s: return "Gyn Oncologist"
        return "Other"

    region = gen.map_region

    # ── build maps for all strata ─────────────────────────────────────────────
    maps = {
        "IHC3": build_maps(gen.MAPS_IHC3),
        "IHC2": build_maps(gen.MAPS_IHC2),
        "IHCU": build_maps(gen.MAPS_IHCU),
    }

    # collect every numeric column we must export
    needed = set()
    for st in STRATA:
        for ind in INDS:
            for lot in LOTS:
                for s, cols in maps[st][ind][lot].items():
                    for k in ("future", "current"):
                        if cols[k]:
                            needed.add(cols[k])
    # weight columns Q3_25Z_{letter}_{lot}_{stratum}
    for ind in INDS:
        for lot in LOTS:
            for st in STRATA:
                needed.add(f"Q3_25Z_{IND_LETTER[ind]}_{lot}_{st}")

    present = [c for c in needed if c in hidx]
    missing = sorted(c for c in needed if c not in hidx)
    if missing:
        print(f"  WARNING: {len(missing)} mapped columns absent from RD (left as 0):")
        for m in missing[:30]:
            print("   ", m)

    # ── build in-memory per-respondent records ────────────────────────────────
    s50 = hidx.get("S0_50Z"); s10 = hidx.get("S0_10Z"); s130 = hidx.get("S0_130Z")
    num_cols = sorted(present)
    records = []
    for i, r in enumerate(data):
        rec = {
            "RespID": i + 1,
            "OL": 1 if i in ol_ids else 0,
            "Practice Setting": practice(r[s50]) if s50 is not None else "Other",
            "Specialty": specialty(r[s10]) if s10 is not None else "Other",
            "Region": region(r[s130]) if s130 is not None else "Other",
        }
        for c in num_cols:
            v = r[hidx[c]]
            rec[c] = v if isinstance(v, (int, float)) else (v if v == "---" else 0)
        records.append(rec)

    # ── microdata xlsx -> data/ (GIT-IGNORED, local validation only, NEVER published)
    data_dir = os.path.join(HERE, "data")
    os.makedirs(data_dir, exist_ok=True)
    out_wb = openpyxl.Workbook(); out_ws = out_wb.active; out_ws.title = "Shares_RD"
    seg_cols = ["RespID", "OL", "Practice Setting", "Specialty", "Region"]
    out_ws.append(seg_cols + num_cols)
    for rec in records:
        out_ws.append([rec[c] for c in seg_cols] + [rec[c] for c in num_cols])
    out_wb.save(os.path.join(data_dir, "Shares_RD.xlsx"))

    # ── aggregate-only shares (NO microdata) -> public/shares_agg.json ─────────
    series_all = gen.ALL_SERIES

    def _median(a):
        s = sorted(a); n = len(s)
        if n == 0: return 0
        m = n // 2
        return s[m] if n % 2 else (s[m - 1] + s[m]) / 2

    def _pct(a, p):
        s = sorted(a)
        if not s: return 0
        idx = (len(s) - 1) * p; lo = int(idx); hi = lo + (1 if idx > lo else 0)
        if lo == hi: return s[lo]
        return s[lo] + (s[hi] - s[lo]) * (idx - lo)

    def cap_weights(weights, method):
        if method == "none": return weights
        pos = [w for w in weights if w > 0]
        if len(pos) < 2: return weights
        cap = _pct(pos, 0.95) if method == "winsor95" else 3 * _median(pos)
        return [cap if w > cap else w for w in weights]

    def cell(rws, ind, lot, st, tf, cap):
        sl = series_all[ind]; mp = maps[st][ind][lot]
        wcol = f"Q3_25Z_{IND_LETTER[ind]}_{lot}_{st}"
        weights, shares = [], []
        for r in rws:
            dv = {}; shown = False
            for s in sl:
                col = mp.get(s, {}).get(tf)
                v = r.get(col) if col else None
                if v in (None, "", "---") or not isinstance(v, (int, float)):
                    dv[s] = 0.0
                else:
                    dv[s] = v / 100.0; shown = True
            if not shown: continue
            w = r.get(wcol); w = w if isinstance(w, (int, float)) else 0
            weights.append(w); shares.append(dv)
        n = len(shares)
        cw = cap_weights(weights, cap); tw = sum(cw) or 1
        wt = {s: sum(cw[i] * shares[i][s] for i in range(n)) / tw for s in sl}
        uw = {s: (sum(shares[i][s] for i in range(n)) / n if n else 0) for s in sl}
        return wt, uw, n, sum(cw)

    def overall(rws, ind, lot, tf, cap):
        sl = series_all[ind]
        cells = {st: cell(rws, ind, lot, st, tf, cap) for st in STRATA}
        vols = {st: cells[st][3] for st in STRATA}
        tot = sum(vols.values()) or 1
        wt = {s: sum(cells[st][0][s] * vols[st] for st in STRATA) / tot for s in sl}
        uw = {s: sum(cells[st][1][s] for st in STRATA) / 3 for s in sl}
        n = max(cells[st][2] for st in STRATA)
        return wt, uw, n

    CAPS = ["none", "winsor95", "median3x"]
    TFS = ["current", "future"]
    STRATA_KEYS = ["Overall"] + STRATA

    # selection menu: All + each single segment value
    selections = [("All", "All respondents", lambda r: True)]
    for col in ("Practice Setting", "Specialty", "Region"):
        for val in sorted({rec[col] for rec in records}):
            selections.append((f"{col}={val}", f"{col}: {val}",
                               (lambda c, v: (lambda r: r[c] == v))(col, val)))

    def cell_block(rws, ind, st, tf):
        out = {"n": {}, "weighted": {c: {} for c in CAPS}, "unweighted": {}}
        for lot in LOTS:
            if st == "Overall":
                for c in CAPS:
                    wt, uw, n = overall(rws, ind, lot, tf, c)
                    for s in series_all[ind]:
                        out["weighted"][c].setdefault(s, []).append(round(wt[s] * 100, 4))
                    if c == "none":
                        for s in series_all[ind]:
                            out["unweighted"].setdefault(s, []).append(round(uw[s] * 100, 4))
                        out["n"][lot] = n
            else:
                for c in CAPS:
                    wt, uw, n, _ = cell(rws, ind, lot, st, tf, c)
                    for s in series_all[ind]:
                        out["weighted"][c].setdefault(s, []).append(round(wt[s] * 100, 4))
                    if c == "none":
                        for s in series_all[ind]:
                            out["unweighted"].setdefault(s, []).append(round(uw[s] * 100, 4))
                        out["n"][lot] = n
        return out

    agg = {"meta": {}, "data": {}}
    sel_meta = []
    for key, label, pred in selections:
        rws_all = [r for r in records if pred(r)]
        rws = [r for r in rws_all if r["OL"] != 1]
        sel_meta.append({"key": key, "label": label, "nActive": len(rws)})
        block = {}
        for tf in TFS:
            block[tf] = {}
            for ind in INDS:
                block[tf][ind] = {}
                for st in STRATA_KEYS:
                    block[tf][ind][st] = cell_block(rws, ind, st, tf)
        agg["data"][key] = block

    agg["meta"] = {
        "indications": {"EC": "Endometrial Cancer", "OC": "Ovarian Cancer", "CC": "Cervical Cancer"},
        "strata": STRATA_KEYS,
        "strataLabels": {"Overall": "Overall", "IHC3": "IHC3+", "IHC2": "IHC2+", "IHCU": "IHC Unknown"},
        "lots": LOTS,
        "timeframes": {"current": "Current (Q3.40)", "future": "Future (Q6.10)"},
        "capMethods": {"none": "Deck (no cap)", "winsor95": "Winsorize @95th", "median3x": "Cap @3× median"},
        "series": series_all,
        "selections": sel_meta,
        "nTotal": len(records),
        "nActive": len([r for r in records if r["OL"] != 1]),
        "builtFrom": os.path.basename(EXCEL_FILE),
    }

    pub = os.path.join(HERE, "public")
    os.makedirs(pub, exist_ok=True)
    agg_path = os.path.join(pub, "shares_agg.json")
    with open(agg_path, "w", encoding="utf-8") as f:
        json.dump(agg, f, separators=(",", ":"))

    # remove any previously-published microdata file
    stale = os.path.join(pub, "Shares_RD.xlsx")
    if os.path.exists(stale):
        os.remove(stale); print(f"  removed stale microdata: {stale}")

    print(f"\nDone (aggregates only — no microdata published).")
    print(f"  respondents: {len(records)}  (OL: {len(ol_ids)}, active: {len(records)-len(ol_ids)})")
    print(f"  selections aggregated: {len(selections)}")
    print(f"  -> {agg_path}  ({os.path.getsize(agg_path)//1024} KB)")
    print(f"  -> {os.path.join(data_dir, 'Shares_RD.xlsx')}  (git-ignored, local only)")


if __name__ == "__main__":
    main()
