"""Validate: the local microdata (data/Shares_RD.xlsx) + local maps (shares_maps.py),
run through a Python clone of the dashboard's share math, must match the deck's own
generate_slides functions (future, weighted) for every indication / stratum / LoT.
Optional one-time correctness check — NOT part of build/deploy. Imports the deck only
here, for comparison."""
import os, importlib.util
import openpyxl
import shares_maps as M

HERE = os.path.dirname(os.path.abspath(__file__))
PARENT = os.path.dirname(HERE)

spec = importlib.util.spec_from_file_location("gen", os.path.join(PARENT, "generate_slides.py"))
gen = importlib.util.module_from_spec(spec); spec.loader.exec_module(gen)

MAPS = {"IHC3": M.MAPS_IHC3, "IHC2": M.MAPS_IHC2, "IHCU": M.MAPS_IHCU}
INDS = ["EC", "OC", "CC"]; LOTS = ["1L", "2L", "3L"]; STRATA = ["IHC3", "IHC2", "IHCU"]

wb = openpyxl.load_workbook(os.path.join(HERE, "data", "Shares_RD.xlsx"), read_only=True, data_only=True)
ws = wb.active
hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
rows = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]
active = [r for r in rows if r.get("OL") != 1]


def js_cell(rws, ind, lot, st):
    series = M.ALL_SERIES[ind]
    mp = MAPS[st][ind].get(lot, {})
    wcol = f"Q3_25Z_{M.IND_LETTER[ind]}_{lot}_{st}"
    weights, shares = [], []
    for r in rws:
        dv, shown = {}, False
        for s in series:
            col = mp.get(s)
            v = r.get(col) if col else None
            if v in (None, "", "---") or not isinstance(v, (int, float)):
                dv[s] = 0.0
            else:
                dv[s] = v / 100.0; shown = True
        if not shown:
            continue
        w = r.get(wcol); w = w if isinstance(w, (int, float)) else 0
        weights.append(w); shares.append(dv)
    n = len(shares); tw = sum(weights) or 1
    wt = {s: sum(weights[i]*shares[i][s] for i in range(n))/tw for s in series}
    return wt, sum(weights)


def js_overall(rws, ind, lot):
    series = M.ALL_SERIES[ind]
    cells = {st: js_cell(rws, ind, lot, st) for st in STRATA}
    vols = {st: cells[st][1] for st in STRATA}
    tot = sum(vols.values()) or 1
    return {s: sum(cells[st][0][s]*vols[st] for st in STRATA)/tot for s in series}


gd = gen.load_data()
data, headers, ol_ids = gd["raw_data"], gd["headers"], gd["ol_ids"]

maxerr = 0.0
for ind in INDS:
    for lot in LOTS:
        strata = gen.compute_strata(data, headers, ol_ids, ind, lot)
        for st in STRATA:
            deck_w = strata[st][0]; js_w = js_cell(active, ind, lot, st)[0]
            for s in M.ALL_SERIES[ind]:
                maxerr = max(maxerr, abs(deck_w.get(s, 0) - js_w.get(s, 0)))
        vols = gen.ihc_patient_volumes(data, headers, ol_ids, ind, lot)
        deck_o, _, _ = gen.overall_shares(strata, M.ALL_SERIES[ind], vols)
        js_o = js_overall(active, ind, lot)
        for s in M.ALL_SERIES[ind]:
            maxerr = max(maxerr, abs(deck_o.get(s, 0) - js_o.get(s, 0)))

print(f"max abs diff (share fraction) clone vs deck: {maxerr:.2e}")
print("PASS" if maxerr < 1e-9 else "FAIL — investigate")
